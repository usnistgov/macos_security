# mscp/generate/excel.py

# Standard python modules
import logging

from pathlib import Path

# Additional python modules
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Local python modules
from src.mscp.common_utils.mobile_config_fix import format_mobileconfig_fix

# Initialize local logger
logger = logging.getLogger(__name__)

def generate_excel(file_out: Path, dataframe: pd.DataFrame) -> None:
    """
    Generate a formatted Excel file from a given DataFrame.

    This function processes the input DataFrame to align with a predefined structure,
    formats its content for better readability in Excel, and writes the processed data
    to an Excel file. It includes modifications to column names, content formatting,
    and layout enhancements like column widths and header alignment.

    Args:
        dataframe (pd.DataFrame): The input DataFrame containing data to be written to Excel.
        file_out (Path): The output file path where the Excel file will be saved.

    DataFrame Processing:
        - Adds and renames columns based on a mapping (`rename_mapping`).
        - Converts nested structures like lists and dictionaries into readable strings.
        - Handles missing data by filling in default values or empty strings.
        - Applies specific formatting for key columns like "Fix" and "Check".
        - Enforces required columns and their order in the output file.

    Excel Formatting:
        - Sets column widths for better readability.
        - Applies alignment and text-wrapping for specific columns.
        - Freezes panes to keep headers visible during scrolling.
        - Formats headers with bold fonts and aligns cell content vertically to the top.

    Example Usage:
        >>> import pandas as pd
        >>> from pathlib import Path
        >>> data = {
        ...     "rule_id": ["Rule1", "Rule2"],
        ...     "title": ["Title1", "Title2"],
        ...     "mobileconfig_info": [{"key1": "value1"}, {}],
        ... }
        >>> df = pd.DataFrame(data)
        >>> generate_excel(df, Path("output.xlsx"))

    Raises:
        KeyError: If critical columns required for processing are missing from the DataFrame.
        ValueError: If `dataframe` contains invalid data types that cannot be processed.

    Notes:
        - The function uses `format_mobileconfig_fix` to format the "Fix" column if `mobileconfig_info` is present.
        - The Excel file is written using `openpyxl` as the engine for `pd.ExcelWriter`.
        - Ensure all required dependencies, like `pandas` and `openpyxl`, are installed.

    Returns:
        None: The function saves the output directly to the specified `file_out` path.
    """

    def __replace_fix(row):
        if row["mobileconfig_info"]:
            try:
                return format_mobileconfig_fix(row["mobileconfig_info"])
            except Exception as e:
                logger.error(f"Error formatting mobileconfig_info: {e}")
                return row["fix"]

        else:
            return row["fix"]

    rename_mapping = {
        "title": "Title",
        "rule_id": "Rule ID",
        "severity": "Severity",
        "discussion": "Discussion",
        "mechanism": "Mechanism",
        "check": "Check",
        "fix": "Fix",
        "cci": "CCI",
        "cce": "CCE",
        "nist_controls": "800-53r5",
        "nist_171": "800-171",
        "disa_stig": "DISA STIG",
        "srg": "SRG",
        "sfr": "SFR",
        "cmmc": "CMMC",
        "indigo": "indigo",
        "custom_refs": "Custom References",
        "tags": "Tags",
        "result": "Check Result",
        "customized": "Modified Rule",
        "benchmark": "CIS Benchmark",
        "controls_v8": "CIS Controls v8"
    }

    list_columns = [
        "cci",
        "cce",
        "nist_controls",
        "nist_171",
        "disa_stig",
        "srg",
        "sfr",
        "cmmc",
        "indigo",
        "custom_refs",
        "odv",
        "tags",
        "benchmark",
        "controls_v8"
    ]

    required_columns = [
        "CCE",
        "Rule ID",
        "Title",
        "Discussion",
        "Mechanism",
        "Check",
        "Check Result",
        "Fix",
        "800-53r5",
        "800-171",
        "SRG",
        "SFR",
        "DISA STIG",
        "CIS Benchmark",
        "CIS Controls v8",
        "CMMC",
        "indigo",
        "CCI",
        "Severity",
        "Modified Rule"
    ]

    # Make a copy of the dataframe so as not to modify the original dataset
    df_copy: pd.DataFrame = dataframe.copy()

    # Dataframe content modifications
    df_copy["section"] = df_copy["section"].astype(pd.CategoricalDtype(ordered=True))

    df_details= df_copy['cis'].apply(lambda x: {} if pd.isna(x) else x).apply(pd.Series)[["benchmark","controls_v8"]]
    df_copy = pd.concat([df_copy, df_details], axis=1)
    df_copy["check"] = df_copy["check"].apply(lambda x: {} if pd.isna(x) else x).apply(pd.Series)
    df_copy["fix"] = df_copy.apply(__replace_fix, axis=1)

    df_copy.columns = (
        df_copy.columns.str.strip()
        .str.strip('[]')
        .str.replace(r"\|", "|", regex=True)
        .str.replace(r"N/A", "", regex=True)
    )

    for col in list_columns:
        if col in df_copy.columns:
            df_copy[col] = df_copy[col].apply(lambda x: "\n".join(x) if isinstance(x, list) else "")

    df_copy = df_copy.drop(columns=["odv","result_value"])
    df_copy.rename(columns=rename_mapping, inplace=True)

    for col in required_columns:
        if col not in df_copy.columns:
            df_copy[col] = ""

    df_copy = df_copy[required_columns]

    with pd.ExcelWriter(file_out, engine='openpyxl') as writer:
        df_copy.to_excel(writer, index=False, header=True, sheet_name='Sheet 1')

        workbook = writer.book
        sheet = writer.sheets["Sheet 1"]
        header_font: Font = Font(bold=True)
        top: Alignment = Alignment(vertical="top")
        topWrap: Alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                    horizontal="left"
                )

        column_widths: dict = {
            0: 15,
            1: 50,
            2: 70,
            3: 95,
            4: 25,
            5: 150,
            6: 25,
            7: 200,
            8: 15,
            9: 15,
            10: 25,
            11: 15,
            12: 15,
            13: 15,
            14: 15,
            15: 15,
            16: 15,
            17: 15,
            18: 15,
            19: 15
        }

        column_alignment: dict = {
            "CCE": top,
            "Rule ID": top,
            "Title": top,
            "Discussion": topWrap,
            "Mechanism": top,
            "Check": topWrap,
            "Check Result": topWrap,
            "Fix": topWrap,
            "800-53r5": topWrap,
            "800-171": topWrap,
            "SRG": topWrap,
            "SFR": topWrap,
            "DISA STIG": topWrap,
            "CIS Benchmark": topWrap,
            "CIS Controls v8": topWrap,
            "CMMC": topWrap,
            "indigo": topWrap,
            "CCI": topWrap,
            "Severity": topWrap,
            "Modified Rule": topWrap
        }

        for cell in sheet[1]:
            cell.font = header_font

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column), start=2):
            for col_idx, cell in enumerate(row, start=1):
                column_name = df_copy.columns[col_idx - 1]
                if column_name in column_alignment:
                    cell.font = Font(size=12)
                    cell.alignment = column_alignment[column_name]


        for col_idx, width in column_widths.items():
            column_letter = get_column_letter(col_idx + 1)
            sheet.column_dimensions[column_letter].width = width

        sheet.freeze_panes = 'C2'
