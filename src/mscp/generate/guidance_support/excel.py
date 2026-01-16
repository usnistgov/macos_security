# mscp/generate/excel.py

# Standard python modules
from pathlib import Path
import json

# Additional python modules
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from ...classes import Baseline
from ...common_utils.logger_instance import logger


def _to_string_item(x):
    """
    Convert an item to a string safely.
    - Strings and scalars become their str() representation.
    - Dicts become compact JSON.
    - Other objects fall back to str().
    """
    if isinstance(x, (str, int, float, bool)):
        return str(x)
    elif isinstance(x, dict):
        # Compact, stable JSON string for dicts
        return json.dumps(x, ensure_ascii=False, sort_keys=True)
    else:
        return str(x)


def _line_display_len(s: str) -> int:
    """
    Returns the display length of a string. If wcwidth is available, uses it for
    better visual width (e.g., emojis / East Asian wide chars). Otherwise, len(s).
    """
    if not s:
        return 0
    return len(s)


def _cell_longest_line_len(cell) -> int:
    """
    Returns the display length of the longest line in the cell.
    - Splits on standard line breaks.
    - Treats non-string values as their string representation.
    - Safely handles None.
    """
    val = cell.value
    if val is None:
        return 0

    try:
        s = str(val)
    except Exception:
        return 0

    # Normalize line endings and split
    lines = s.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    return max((_line_display_len(line) for line in lines), default=0)


def auto_fit_columns(
    sheet, threshold_length: int = 120, buffer: int = 2, wrap_multiline: bool = True
):
    """
    Auto-fit width of each column in `sheet` based on the longest line per cell.
    - threshold_length: max width to set for any column (your cap).
    - buffer: add a few extra characters of padding.
    - wrap_multiline: optionally enable wrapText for cells that contain newlines.
    """
    EXCEL_MAX_COL_WIDTH = 255  # Excel hard limit

    for col_cells in sheet.iter_cols():
        max_len = 0

        for cell in col_cells:
            longest = _cell_longest_line_len(cell)
            if longest > max_len:
                max_len = longest

            # Optional: enable wrapText if the cell actually has newlines
            if wrap_multiline:
                val = cell.value
                if isinstance(val, str) and ("\n" in val or "\r" in val):
                    if cell.alignment is None or not cell.alignment.wrapText:
                        # Keep other alignment attributes if present
                        cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal
                            if cell.alignment
                            else None,
                            vertical=cell.alignment.vertical
                            if cell.alignment
                            else None,
                            wrapText=True,
                        )

        # Column letter (e.g., 'A')
        column_letter = get_column_letter(col_cells[0].column)

        # Compute final width with buffer and caps
        target_width = max_len + buffer
        target_width = min(target_width, threshold_length, EXCEL_MAX_COL_WIDTH)

        sheet.column_dimensions[column_letter].width = target_width


def format_list_cell(x, unwrap_single=True, sep="\n"):
    """
    Format a cell that may be a list:
    - Empty list -> pd.NA
    - Single-item list (unwrap_single=True) -> that item
    - Multi-item list -> newline-separated string (items coerced to string)
    - Non-list values returned unchanged
    """
    if not isinstance(x, list):
        return x
    if len(x) == 0:
        return pd.NA
    if unwrap_single and len(x) == 1:
        return x[0]
    return sep.join(_to_string_item(item) for item in x)


def _ensure_unique_names(names, taken):
    """
    Ensure new column names do not collide with existing ones.
    If collision occurs, append __dupN suffix.
    """
    finalized = []
    for base in names:
        name = base
        i = 1
        while name in taken:
            name = f"{base}__dup{i}"
            i += 1
        finalized.append(name)
        taken.add(name)
    return finalized


def expand_dict_column(
    df,
    col,
    unwrap_single_lists=True,
    list_sep="\n",
    drop_original=True,
    flatten_sep=".",
):
    """
    Expand dictionaries found in df[col] into new columns named '{col}:{keypath}',
    flattening nested keys with `flatten_sep`. Lists encountered inside dict values
    are formatted per `format_list_cell`. Original column can be dropped.
    """
    # Identify rows with dicts
    mask = df[col].apply(lambda v: isinstance(v, dict))

    if not mask.any():
        return df

    # Normalize the dicts (nested keys flattened with flatten_sep)
    expanded = pd.json_normalize(df.loc[mask, col], sep=flatten_sep)

    # Align to full DataFrame index (non-dict rows get NaN)
    expanded = expanded.reindex(df.index)

    # Format lists inside the expanded dict values
    expanded = expanded.map(
        lambda v: format_list_cell(v, unwrap_single_lists, list_sep)
    )

    # Prefix with original column name
    new_names = [f"{col}:{c}" for c in expanded.columns]

    # Avoid collisions with existing columns
    taken = set(df.columns)
    expanded.columns = _ensure_unique_names(new_names, taken)

    # Concatenate expanded columns
    out = pd.concat([df, expanded], axis=1)

    # Optionally drop the original dict column
    if drop_original:
        out = out.drop(columns=[col])

    return out


def expand_dicts_and_format_lists(
    df,
    unwrap_single_lists=True,
    list_sep="\n",
    drop_original_dict_cols=True,
    flatten_sep=".",
):
    """
    Process the entire DataFrame:
      1) For every column that contains dictionaries in any row:
         - Expand into '{col}:{keypath}' columns.
         - Drop the original dict column (configurable).
         - Format lists found inside those dict values.
      2) For every remaining column that contains lists in any row:
         - Unwrap single-item lists.
         - Join multi-item lists with `list_sep`.
    """
    out = df.copy()

    # First pass: expand dict-containing columns
    for col in list(out.columns):
        if out[col].apply(lambda v: isinstance(v, dict)).any():
            out = expand_dict_column(
                out,
                col,
                unwrap_single_lists=unwrap_single_lists,
                list_sep=list_sep,
                drop_original=drop_original_dict_cols,
                flatten_sep=flatten_sep,
            )

    # Second pass: format standalone list-containing columns
    for col in list(out.columns):
        if out[col].apply(lambda v: isinstance(v, list)).any():
            out[col] = out[col].apply(
                lambda v: format_list_cell(v, unwrap_single_lists, list_sep)
            )

    return out


def generate_excel(file_out: Path, baseline: Baseline) -> None:
    """
    Generates an Excel file from the given baseline data.

    Args:
        file_out (Path): The output file path where the Excel file will be saved.
        baseline (Baseline): The baseline data to be converted into an Excel file.

    Returns:
        None

    This function performs the following steps:
    1. Logs the start of the Excel generation process.
    2. Converts the baseline data to a DataFrame and makes a copy of it.
    3. Drops unwanted columns from the DataFrame.
    4. Modifies the DataFrame content, including handling nested structures and renaming columns.
    5. Ensures all required columns are present in the DataFrame, dropping any columns that have all None values.
    6. Writes the DataFrame to an Excel file using the openpyxl engine.
    7. Applies formatting to the Excel sheet, including setting column widths, header fonts, and cell alignments.
    8. Logs the successful generation of the Excel file.

    Raises:
        Any exceptions raised during the process will be logged.
    """
    logger.info("Starting Excel generation process.")

    logger.debug("Converting baseline to DataFrame.")
    dataframe = baseline.to_dataframe()

    # drop unnecessary columns
    dataframe.drop("finding", axis=1, inplace=True)
    dataframe.drop("uuid", axis=1, inplace=True)
    # dataframe.drop("section", axis=1, inplace=True)
    dataframe.drop("platforms", axis=1, inplace=True)
    dataframe.drop("os_name", axis=1, inplace=True)
    dataframe.drop("os_type", axis=1, inplace=True)
    dataframe.drop("os_version", axis=1, inplace=True)
    dataframe.drop("odv", axis=1, inplace=True)
    dataframe.drop("tags", axis=1, inplace=True)
    dataframe.drop("mobileconfig_info", axis=1, inplace=True)
    dataframe.drop("ddm_info", axis=1, inplace=True)

    column_order = [
        "rule_id",
        "title",
        "section",
        "discussion",
        "mechanism",
        "check",
        "result_value",
        "fix",
        "default_state",
        "severity",
        "customized",
        "nist",
        "disa",
        "cis",
        "bsi",
        "custom_refs",
    ]

    sorted_dataframe = dataframe[column_order]

    df2 = expand_dicts_and_format_lists(
        sorted_dataframe,
        unwrap_single_lists=True,  # single-item lists -> element
        list_sep="\n",  # multi-item lists -> newline-separated
        drop_original_dict_cols=True,  # drop dict columns (like CIS) after expansion
        flatten_sep=".",  # nested keys flattened with dots
    )

    # move CCE to front, and customized to back
    cce_column = df2.pop("nist:cce")
    df2.insert(0, "nist:cce", cce_column)

    df2["customized"] = df2.pop("customized")

    # drop any columns that don't have any values
    df2.dropna(axis=1, how="all", inplace=True)

    # convert column headers to uppercase
    df2.columns = df2.columns.str.upper()

    logger.debug("Writing DataFrame to Excel file.")

    with pd.ExcelWriter(file_out, engine="openpyxl") as writer:
        df2.to_excel(writer, index=False, header=True, sheet_name=baseline.name)

        sheet = writer.sheets[baseline.name]
        header_font: Font = Font(bold=True)
        top: Alignment = Alignment(vertical="top")
        topWrap: Alignment = Alignment(
            vertical="top", wrap_text=True, horizontal="left"
        )

        top_columns: list = [
            "title",
            "rule_id",
            "result_value",
            "mechanism",
            "section",
        ]

        for cell in sheet[1]:
            cell.font = header_font

        for _, row in enumerate(
            sheet.iter_rows(
                min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
            ),
            start=2,
        ):
            for col_idx, cell in enumerate(row, start=1):
                column_name = df2.columns[col_idx - 1]
                if column_name in top_columns:
                    cell.alignment = top
                else:
                    cell.alignment = topWrap

        auto_fit_columns(sheet, threshold_length=120, buffer=2, wrap_multiline=True)

        # Create and add a table for easy use
        max_row = sheet.max_row
        max_col = sheet.max_column
        ref_range = f"A1:{get_column_letter(max_col)}{max_row}"

        style = TableStyleInfo(name="TableStyleMedium2")
        table = Table(displayName=baseline.name, ref=ref_range)
        table.tableStyleInfo = style
        sheet.add_table(table)

        sheet.freeze_panes = "C2"

    logger.success("Excel file generated successfully at {}.", file_out)
