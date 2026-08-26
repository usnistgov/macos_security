# mscp/generate/spreadsheet.py
"""Spreadsheet (CSV/XLSX) export and import for bulk rule editing.

Export writes one row per rule with a leading metadata row
(rule_id = ``__meta__``) that carries os_type and os_version so the
importer knows which versioned YAML keys to update.  Import reads
the file back and patches each rule's source YAML in-place using
ruamel.yaml so the rest of the file's formatting is preserved.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Union

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from ruamel.yaml import YAML
from ruamel.yaml.scalarstring import LiteralScalarString

from ..classes.baseline import Baseline
from ..classes.macsecurityrule import Macsecurityrule
from ..classes.rule_library import RuleLibrary
from ..common_utils import config
from ..common_utils.logger_instance import logger

# ── Constants ──────────────────────────────────────────────────────────────────

LIST_SEP = ";"
META_RULE_ID = "__meta__"

FIXED_COLUMNS: list[str] = [
    "rule_id",
    "category",
    "title",
    "discussion",
    "tags",
    "nist_800_53r5",
    "nist_800_171r3",
    "cce",
    "disa_cci",
    "disa_srg",
    "disa_stig",
    "disa_cmmc",
    "cis_benchmark",
    "cis_controls_v8",
    "bzk_bio",
    "bsi_indigo",
    "hhs_hicp",
    "benchmarks",
    "odv_hint",
    "odv_datatype",
    "odv_recommended",
    "odv_custom",
]

META_COLS: list[str] = ["_meta_os_type", "_meta_os_version"]

_ryaml = YAML()
_ryaml.preserve_quotes = True
_ryaml.width = 4096
_ryaml.indent(mapping=2, sequence=4, offset=2)


# ── Shared helpers ─────────────────────────────────────────────────────────────

def _list_to_str(v: Any) -> str:
    if not v:
        return ""
    if isinstance(v, list):
        return LIST_SEP.join(str(x) for x in v if x is not None)
    return str(v)


def _str_to_list(s: str) -> list[str]:
    if not s or not s.strip():
        return []
    return [x.strip() for x in s.split(LIST_SEP) if x.strip()]


def _has_change_list(
    current: list | None, new_str: str
) -> tuple[bool, list[str]]:
    """Return (changed, new_list). changed is False if new_str is empty."""
    if not new_str.strip():
        return False, []
    new_list = _str_to_list(new_str)
    if not new_list:
        return False, []
    return sorted(str(v) for v in (current or [])) != sorted(new_list), new_list


def _gather_odv_keys(rules: list[Macsecurityrule]) -> list[str]:
    """Collect per-benchmark ODV keys present across rules (excluding hint/custom/recommended)."""
    keys: set[str] = set()
    for rule in rules:
        if rule.odv:
            for k in rule.odv:
                if k not in ("hint", "custom", "recommended"):
                    keys.add(k)
    return sorted(keys)


def _find_rule_file(rule_id: str) -> Path | None:
    rules_dirs = [
        Path(config["rules_dir"]),
        Path(config["custom"]["rules_dir"]),
    ]
    return next(
        (
            f
            for rd in rules_dirs
            if rd.exists()
            for f in rd.rglob(f"{rule_id}.y*ml")
        ),
        None,
    )


def _valid_categories() -> set[str]:
    """Category subfolder names that `collect_platform_rules` scans for rules.

    Mirrors that method's own exclusion of ``sysprefs`` so a category
    written here is guaranteed to be picked back up during rule
    collection / baseline generation.
    """
    rules_dir = Path(config["rules_dir"])
    if not rules_dir.exists():
        return set()
    return {p.name for p in rules_dir.iterdir() if p.is_dir() and p.name != "sysprefs"}


def _set_nested(data: Any, keys: list[str], value: Any) -> None:
    """Set ``data[keys[0]][keys[1]]... = value``, creating intermediate dicts as needed.

    Unlike ``dict.get(k, {})``, which returns a throwaway dict that is
    never attached back to its parent when *k* is absent, this walks the
    chain with ``get``/assignment at each level so a namespace that didn't
    previously exist (e.g. a rule with no ``references.bzk`` block) is
    created in place and the write actually lands in *data*.
    """
    node = data
    for key in keys[:-1]:
        nxt = node.get(key)
        if nxt is None:
            nxt = {}
            node[key] = nxt
        node = nxt
    node[keys[-1]] = value


def _coerce_odv(value_str: str, reference: Any) -> Any:
    """Coerce value_str to match the Python type of reference."""
    if isinstance(reference, bool):
        return value_str.lower() in ("true", "1", "yes")
    if isinstance(reference, int):
        try:
            return int(value_str)
        except ValueError:
            return value_str
    if isinstance(reference, float):
        try:
            return float(value_str)
        except ValueError:
            return value_str
    return value_str


def _infer_odv_datatype(row: dict[str, str]) -> str:
    """Best-effort ``odv.hint.datatype`` guess when a row sets a hint description
    but no explicit ``odv_datatype``. Mirrors the ``datatype`` values actually
    used across the rule set (``number`` / ``string`` / ``enum`` / ``regex``) --
    only ``number`` is inferable from a bare value, so anything else falls
    back to ``string``.
    """
    for col in ("odv_recommended", "odv_custom"):
        val = row.get(col, "").strip()
        if not val:
            continue
        try:
            float(val)
            return "number"
        except ValueError:
            pass
    return "string"


# ── Export ─────────────────────────────────────────────────────────────────────

def _rules_from_source(
    source: Union[Baseline, RuleLibrary, Path],
) -> tuple[list[Macsecurityrule], str, float, str]:
    """Return (rules, os_type, os_version, name) from any supported source."""
    if isinstance(source, Path):
        source = Baseline.from_yaml(source)
    if isinstance(source, Baseline):
        rules = [r for p in source.profile for r in p.rules]
        os_type = source.platform.get("os", "macOS")
        os_version = float(source.platform.get("version", 0.0))
        return rules, os_type, os_version, source.name
    if isinstance(source, RuleLibrary):
        rules = list(source)
        if not rules:
            return rules, "macOS", 0.0, ""
        versions = {(r.os_type, float(r.os_version)) for r in rules}
        if len(versions) > 1:
            raise ValueError(
                "RuleLibrary spans multiple platform/version combinations "
                f"({sorted(versions)}); spreadsheet export carries a single "
                "os_type/os_version in its metadata row, so a mixed library "
                "can't be exported as one sheet. Narrow it first with "
                "by_platform() and/or by_os()."
            )
        os_type, os_version = next(iter(versions))
        return rules, os_type, os_version, ""
    raise TypeError(f"Unsupported source type: {type(source)}")


def _rule_to_row(rule: Macsecurityrule, odv_keys: list[str]) -> dict[str, str]:
    """Flatten one rule into a flat string dict suitable for a spreadsheet row."""
    import yaml as _pyyaml

    os_version_str = str(float(rule.os_version))
    refs = rule.references
    nist = refs.nist
    disa = refs.disa
    cis = refs.cis
    bzk = refs.bzk
    bsi = refs.bsi
    hhs = refs.hhs

    # Category subfolder (e.g. "system_settings") the rule lives under in
    # rules_dir — this is where a brand-new rule must be created too, since
    # that's what `collect_platform_rules` scans for rule discovery.
    category = rule.source_file.parent.name if rule.source_file else ""

    # Read title/discussion directly from the source YAML to preserve any
    # $ODV placeholders — the loaded rule object has them already resolved.
    raw_title = rule.title or ""
    raw_discussion = rule.discussion or ""
    if rule.source_file and rule.source_file.is_file():
        try:
            raw = _pyyaml.safe_load(rule.source_file.read_text(encoding="utf-8"))
            raw_title = raw.get("title", raw_title)
            raw_discussion = raw.get("discussion", raw_discussion)
        except Exception:
            pass

    bm_list: list[str] = []
    for bm in (
        rule.platforms.get(rule.os_type, {})
        .get(os_version_str, {})
        .get("benchmarks", [])
    ):
        name = bm.get("name", "")
        if not name:
            continue
        sev = bm.get("severity", "")
        bm_list.append(f"{name}:{sev}" if sev else name)

    odv = rule.odv or {}
    hint = odv.get("hint", {})
    hint_desc = hint.get("description", "") if isinstance(hint, dict) else str(hint)
    hint_datatype = hint.get("datatype", "") if isinstance(hint, dict) else ""

    row: dict[str, str] = {
        "rule_id": rule.rule_id,
        "category": category,
        "title": raw_title,
        "discussion": raw_discussion,
        "tags": _list_to_str(rule.tags),
        "nist_800_53r5": _list_to_str(nist.nist_800_53r5 if nist else None),
        "nist_800_171r3": _list_to_str(nist.nist_800_171r3 if nist else None),
        "cce": _list_to_str(nist.cce if nist else None),
        "disa_cci": _list_to_str(disa.cci if disa else None),
        "disa_srg": _list_to_str(disa.srg if disa else None),
        "disa_stig": _list_to_str(disa.disa_stig if disa else None),
        "disa_cmmc": _list_to_str(disa.cmmc if disa else None),
        "cis_benchmark": _list_to_str(cis.benchmark if cis else None),
        "cis_controls_v8": _list_to_str(cis.controls_v8 if cis else None),
        "bzk_bio": _list_to_str(bzk.bio if bzk else None),
        "bsi_indigo": _list_to_str(bsi.indigo if bsi else None),
        "hhs_hicp": _list_to_str(hhs.hicp if hhs else None),
        "benchmarks": _list_to_str(bm_list),
        "odv_hint": hint_desc,
        "odv_datatype": hint_datatype,
        "odv_recommended": str(odv.get("recommended", "")),
        "odv_custom": str(odv.get("custom", "")),
        "_meta_os_type": "",
        "_meta_os_version": "",
    }

    for k in odv_keys:
        row[f"odv_{k}"] = str(odv.get(k, ""))

    return row


def _build_columns(odv_keys: list[str]) -> list[str]:
    return FIXED_COLUMNS + [f"odv_{k}" for k in odv_keys] + META_COLS


def _write_csv(
    output_path: Path,
    columns: list[str],
    meta_row: dict[str, str],
    data_rows: list[dict[str, str]],
) -> None:
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerow(meta_row)
        writer.writerows(data_rows)


def _write_xlsx(
    output_path: Path,
    columns: list[str],
    meta_row: dict[str, str],
    data_rows: list[dict[str, str]],
    sheet_name: str,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name or "Rules"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    meta_fill = PatternFill("solid", fgColor="F2F2F2")

    # Header row
    ws.append(columns)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Metadata row (hidden)
    ws.append([meta_row.get(c, "") for c in columns])
    for cell in ws[2]:
        cell.fill = meta_fill
    ws.row_dimensions[2].hidden = True

    # Data rows
    for row in data_rows:
        ws.append([row.get(c, "") for c in columns])

    # Hide metadata columns
    for idx, col in enumerate(columns, 1):
        if col.startswith("_meta_"):
            ws.column_dimensions[get_column_letter(idx)].hidden = True

    # Auto-fit widths (capped)
    for col_cells in ws.iter_cols():
        max_len = max(
            (len(str(cell.value or "").split("\n")[0]) for cell in col_cells),
            default=0,
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

    ws.freeze_panes = "B3"

    wb.save(output_path)


def export_spreadsheet(
    source: Union[Baseline, RuleLibrary, Path],
    output_path: Path,
) -> None:
    """Export rules to a CSV or XLSX file for bulk editing.

    The first data row (rule_id = ``__meta__``) carries os_type and
    os_version so the importer (``mscp admin import``) can route
    version-specific updates to the correct YAML key
    (e.g. ``references.nist.cce.macos_15``).

    Args:
        source: A ``Baseline``, ``RuleLibrary``, or path to a baseline YAML.
        output_path: Destination ``.csv`` or ``.xlsx`` file.
    """
    rules, os_type, os_version, name = _rules_from_source(source)
    odv_keys = _gather_odv_keys(rules)
    columns = _build_columns(odv_keys)

    meta_row: dict[str, str] = {c: "" for c in columns}
    meta_row["rule_id"] = META_RULE_ID
    meta_row["_meta_os_type"] = os_type
    meta_row["_meta_os_version"] = str(os_version)

    data_rows = [_rule_to_row(rule, odv_keys) for rule in rules]

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if output_path.suffix.lower() in (".xlsx", ".xls"):
        _write_xlsx(output_path, columns, meta_row, data_rows, name)
    else:
        _write_csv(output_path, columns, meta_row, data_rows)

    logger.success("Exported {} rules to {}.", len(data_rows), output_path)
    print(f"Exported {len(data_rows)} rules to {output_path}")


def export_template(os_type: str, os_version: float, output_path: Path) -> None:
    """Export a blank rule template CSV or XLSX.

    Produces just the header row and the metadata row so the user can
    fill in new rule data and import it with ``mscp admin import``.

    Args:
        os_type: Target OS type string (e.g. ``"macOS"``).
        os_version: Target OS version float (e.g. ``26.0``).
        output_path: Destination ``.csv`` or ``.xlsx`` file.
    """
    columns = _build_columns([])  # fixed columns only; no dynamic ODV keys

    meta_row: dict[str, str] = {c: "" for c in columns}
    meta_row["rule_id"] = META_RULE_ID
    meta_row["_meta_os_type"] = os_type
    meta_row["_meta_os_version"] = str(os_version)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if output_path.suffix.lower() in (".xlsx", ".xls"):
        _write_xlsx(output_path, columns, meta_row, [], "template")
    else:
        _write_csv(output_path, columns, meta_row, [])

    logger.success("Exported rule template to {}.", output_path)
    print(f"Exported rule template to {output_path}")


# ── Import ─────────────────────────────────────────────────────────────────────

def _read_rows(
    input_path: Path,
) -> tuple[dict[str, str], list[dict[str, str]]]:
    """Return (meta_dict, data_rows) from a CSV or XLSX file."""
    if input_path.suffix.lower() in (".xlsx", ".xls"):
        return _read_xlsx_rows(input_path)
    return _read_csv_rows(input_path)


def _read_csv_rows(
    input_path: Path,
) -> tuple[dict[str, str], list[dict[str, str]]]:
    with input_path.open("r", newline="", encoding="utf-8-sig") as f:
        rows = [dict(r) for r in csv.DictReader(f)]

    meta: dict[str, str] = {}
    data: list[dict[str, str]] = []
    for row in rows:
        if row.get("rule_id", "").strip() == META_RULE_ID:
            meta = row
        else:
            data.append(row)
    return meta, data


def _read_xlsx_rows(
    input_path: Path,
) -> tuple[dict[str, str], list[dict[str, str]]]:
    wb = openpyxl.load_workbook(input_path)
    ws = next(
        (wb[s] for s in wb.sheetnames if s.lower() != "meta"),
        wb.active,
    )

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return {}, []

    headers = [str(h) if h is not None else "" for h in all_rows[0]]

    meta: dict[str, str] = {}
    data: list[dict[str, str]] = []
    for row_values in all_rows[1:]:
        row = {
            headers[i]: str(v) if v is not None else ""
            for i, v in enumerate(row_values)
            if i < len(headers)
        }
        if row.get("rule_id", "").strip() == META_RULE_ID:
            meta = row
        else:
            data.append(row)

    return meta, data


def _apply_row_to_yaml(
    data: Any,
    row: dict[str, str],
    os_type: str,
    os_version: float,
    os_typeversion: str,
) -> list[str]:
    """Patch the ruamel.yaml-loaded *data* dict with values from *row*.

    Only writes a field when the spreadsheet value is non-empty AND
    differs from the current YAML value.  Returns a list of field names
    that were changed (empty list = no changes).
    """
    changed_fields: list[str] = []
    os_version_str = str(float(os_version))

    # ── simple string fields ──────────────────────────────────────────────────
    new_title = row.get("title", "").strip()
    if new_title and new_title != str(data.get("title", "")).strip():
        data["title"] = new_title
        changed_fields.append("title")

    new_disc = row.get("discussion", "")
    if new_disc.strip() and new_disc.strip() != str(data.get("discussion", "")).strip():
        data["discussion"] = (
            LiteralScalarString(new_disc) if "\n" in new_disc else new_disc
        )
        changed_fields.append("discussion")

    # ── tags ──────────────────────────────────────────────────────────────────
    changed, new_list = _has_change_list(data.get("tags"), row.get("tags", ""))
    if changed:
        data["tags"] = new_list
        changed_fields.append("tags")

    # ── references ────────────────────────────────────────────────────────────
    # NOTE: namespaces are looked up by name (not pre-fetched into a dict
    # variable) because a namespace absent from this rule's YAML must be
    # created and attached back into `data` at write time -- `refs.get(ns)
    # or {}` alone would silently mutate a throwaway dict that's never
    # linked to `data`, so the write is dropped even though it's reported
    # as a change. See `_set_nested`.
    refs = data.get("references", {})

    # Simple (non-versioned) list references
    simple_ref_map = [
        ("nist", "800-53r5",   "nist_800_53r5"),
        ("nist", "800-171r3",  "nist_800_171r3"),
        ("disa", "cci",        "disa_cci"),
        ("disa", "srg",        "disa_srg"),
        ("disa", "cmmc",       "disa_cmmc"),
        ("bzk",  "bio",        "bzk_bio"),
        ("bsi",  "indigo",     "bsi_indigo"),
        ("hhs",  "hicp",       "hhs_hicp"),
    ]
    for namespace, yaml_key, col in simple_ref_map:
        current = (refs.get(namespace) or {}).get(yaml_key) if refs else None
        changed, new_list = _has_change_list(current, row.get(col, ""))
        if changed:
            _set_nested(data, ["references", namespace, yaml_key], sorted(new_list))
            changed_fields.append(col)

    # cis_controls_v8 (list of floats, special handling)
    new_str = row.get("cis_controls_v8", "")
    if new_str.strip():
        try:
            new_floats = sorted(float(x) for x in _str_to_list(new_str))
        except ValueError:
            new_floats = []
        if new_floats:
            cis = (refs.get("cis") or {}) if refs else {}
            current = sorted(float(x) for x in (cis.get("controls_v8") or []))
            if current != new_floats:
                _set_nested(data, ["references", "cis", "controls_v8"], new_floats)
                changed_fields.append("cis_controls_v8")

    # Version-specific list references
    versioned_ref_map = [
        ("nist", "cce",       "cce"),
        ("disa", "disa_stig", "disa_stig"),
        ("cis",  "benchmark", "cis_benchmark"),
    ]
    for namespace, yaml_key, col in versioned_ref_map:
        new_str = row.get(col, "")
        if not new_str.strip():
            continue
        new_list = _str_to_list(new_str)
        ns_dict = (refs.get(namespace) or {}) if refs else {}
        field_val = ns_dict.get(yaml_key, {})
        if isinstance(field_val, dict):
            # Version-keyed dict: only update the target version
            current = field_val.get(os_typeversion) or []
            if sorted(str(v) for v in current) != sorted(new_list):
                _set_nested(
                    data,
                    ["references", namespace, yaml_key, os_typeversion],
                    sorted(new_list),
                )
                changed_fields.append(col)
        else:
            # Direct list (not version-keyed in this file)
            current = field_val or []
            if sorted(str(v) for v in current) != sorted(new_list):
                _set_nested(data, ["references", namespace, yaml_key], sorted(new_list))
                changed_fields.append(col)

    # ── platforms benchmarks ──────────────────────────────────────────────────
    new_bm_str = row.get("benchmarks", "")
    if new_bm_str.strip():
        new_bm_dicts: list[dict] = []
        for entry in _str_to_list(new_bm_str):
            if ":" in entry:
                name, sev = entry.rsplit(":", 1)
                new_bm_dicts.append({"name": name.strip(), "severity": sev.strip()})
            else:
                new_bm_dicts.append({"name": entry.strip()})

        platforms = data.get("platforms", {})
        if platforms and os_type in platforms and os_version_str in platforms[os_type]:
            current_bm = platforms[os_type][os_version_str].get("benchmarks", [])

            def _norm(b: Any) -> tuple[str, str]:
                return (str(b.get("name", "")), str(b.get("severity", "")))

            if sorted(_norm(b) for b in current_bm) != sorted(_norm(b) for b in new_bm_dicts):
                platforms[os_type][os_version_str]["benchmarks"] = new_bm_dicts
                changed_fields.append("benchmarks")

    # ── ODV fields ────────────────────────────────────────────────────────────
    if "odv" in data:
        odv = data["odv"]

        # hint.description / hint.datatype: merged into the existing hint
        # dict (rather than written as `odv["hint"]`/`odv["datatype"]`
        # directly) so schema fields this sheet doesn't expose --
        # `validation` -- survive, and so `odv.hint` keeps the
        # {datatype, description, validation} shape `OdvHint` and
        # `odv_query`'s `odv_hint['description']` access expect.
        new_hint_desc = row.get("odv_hint", "").strip()
        new_hint_datatype = row.get("odv_datatype", "").strip()
        if new_hint_desc or new_hint_datatype:
            hint = odv.get("hint")
            if not isinstance(hint, dict):
                hint = {}
            if new_hint_desc and new_hint_desc != hint.get("description", ""):
                hint["description"] = new_hint_desc
                changed_fields.append("odv_hint")
            if new_hint_datatype and new_hint_datatype != hint.get("datatype", ""):
                hint["datatype"] = new_hint_datatype
                changed_fields.append("odv_datatype")
            odv["hint"] = hint

        odv_cols = [
            (col, col[4:])
            for col in row
            if col.startswith("odv_") and col not in ("odv_hint", "odv_datatype")
        ]
        for col, key in odv_cols:
            new_str = row.get(col, "").strip()
            if new_str:
                current = odv.get(key)
                coerced = _coerce_odv(new_str, current)
                if coerced != current:
                    odv[key] = coerced
                    changed_fields.append(col)

    return changed_fields


def _create_rule_yaml(
    row: dict[str, str],
    os_type: str,
    os_version: float,
    os_typeversion: str,
) -> dict:
    """Build a minimal rule YAML dict from a spreadsheet row."""
    rule_id = row.get("rule_id", "").strip()
    os_version_str = str(float(os_version))

    # ── references.nist ──────────────────────────────────────────────────────
    nist: dict = {}
    n800_53r5 = _str_to_list(row.get("nist_800_53r5", ""))
    if n800_53r5:
        nist["800-53r5"] = sorted(n800_53r5)
    n800_171r3 = _str_to_list(row.get("nist_800_171r3", ""))
    if n800_171r3:
        nist["800-171r3"] = sorted(n800_171r3)
    cce = _str_to_list(row.get("cce", ""))
    if cce:
        nist["cce"] = {os_typeversion: sorted(cce)}

    refs: dict = {"nist": nist}

    # ── references.disa ──────────────────────────────────────────────────────
    disa_cci = _str_to_list(row.get("disa_cci", ""))
    disa_srg = _str_to_list(row.get("disa_srg", ""))
    disa_stig = _str_to_list(row.get("disa_stig", ""))
    disa_cmmc = _str_to_list(row.get("disa_cmmc", ""))
    if any([disa_cci, disa_srg, disa_stig, disa_cmmc]):
        disa: dict = {}
        if disa_cci:
            disa["cci"] = sorted(disa_cci)
        if disa_srg:
            disa["srg"] = sorted(disa_srg)
        if disa_stig:
            disa["disa_stig"] = {os_typeversion: sorted(disa_stig)}
        if disa_cmmc:
            disa["cmmc"] = sorted(disa_cmmc)
        refs["disa"] = disa

    # ── references.cis ───────────────────────────────────────────────────────
    cis_bm = _str_to_list(row.get("cis_benchmark", ""))
    cis_cv8_str = row.get("cis_controls_v8", "").strip()
    if cis_bm or cis_cv8_str:
        cis: dict = {}
        if cis_bm:
            cis["benchmark"] = {os_typeversion: sorted(cis_bm)}
        if cis_cv8_str:
            try:
                cis["controls_v8"] = sorted(float(x) for x in _str_to_list(cis_cv8_str))
            except ValueError:
                pass
        refs["cis"] = cis

    if (bzk := _str_to_list(row.get("bzk_bio", ""))):
        refs["bzk"] = {"bio": sorted(bzk)}
    if (bsi := _str_to_list(row.get("bsi_indigo", ""))):
        refs["bsi"] = {"indigo": sorted(bsi)}
    if (hhs := _str_to_list(row.get("hhs_hicp", ""))):
        refs["hhs"] = {"hicp": sorted(hhs)}

    # ── platforms ────────────────────────────────────────────────────────────
    bm_dicts: list[dict] = []
    for entry in _str_to_list(row.get("benchmarks", "")):
        if ":" in entry:
            name, sev = entry.rsplit(":", 1)
            bm_dicts.append({"name": name.strip(), "severity": sev.strip()})
        else:
            bm_dicts.append({"name": entry.strip()})

    # ── ODV ──────────────────────────────────────────────────────────────────
    odv: dict = {}
    odv_hint_desc = row.get("odv_hint", "").strip()
    odv_hint_datatype = row.get("odv_datatype", "").strip()
    if odv_hint_desc or odv_hint_datatype:
        # `OdvHint` requires both `datatype` and `description` -- fill in
        # a best-effort datatype when the sheet only set the description,
        # so the rule doesn't crash `Macsecurityrule.validate_odv_hint` /
        # `odv_query`'s `odv_hint['description']` access later.
        odv["hint"] = {
            "datatype": odv_hint_datatype or _infer_odv_datatype(row),
            "description": odv_hint_desc,
        }
    odv_rec = row.get("odv_recommended", "").strip()
    if odv_rec:
        odv["recommended"] = _coerce_odv(odv_rec, None)
    odv_cust = row.get("odv_custom", "").strip()
    if odv_cust:
        odv["custom"] = _coerce_odv(odv_cust, None)
    for col in row:
        if col.startswith("odv_") and col not in (
            "odv_hint", "odv_datatype", "odv_recommended", "odv_custom",
        ):
            val = row[col].strip()
            if val:
                odv[col[4:]] = _coerce_odv(val, None)

    # ── assemble ─────────────────────────────────────────────────────────────
    discussion = row.get("discussion", "").strip()
    yaml_dict: dict = {
        "id": rule_id,
        "title": row.get("title", "").strip(),
        "discussion": LiteralScalarString(discussion + "\n") if "\n" in discussion else discussion,
        "references": refs,
        "platforms": {
            os_type: {
                os_version_str: {
                    "benchmarks": bm_dicts,
                }
            }
        },
    }
    if (tags := _str_to_list(row.get("tags", ""))):
        yaml_dict["tags"] = tags
    if odv:
        yaml_dict["odv"] = odv

    return yaml_dict


def import_spreadsheet(
    input_path: Path,
    dry_run: bool = False,
) -> dict[str, int]:
    """Read *input_path* and patch source rule YAML files in-place.

    Each rule row is matched to its source YAML via rule_id, loaded with
    ruamel.yaml (preserving formatting), updated in place, and written
    back.  Version-specific fields (CCE, DISA STIG, CIS benchmark,
    benchmarks list) are routed to the correct OS-version key using the
    os_type / os_version from the metadata row.

    Prints a formatted per-rule report to stdout regardless of log level.

    Args:
        input_path: Path to the exported CSV or XLSX.
        dry_run: Compute diffs but do not write any files.

    Returns:
        Summary dict with keys ``modified``, ``skipped``, ``not_found``,
        ``created``.
    """
    meta, rows = _read_rows(input_path)

    os_type = meta.get("_meta_os_type", "").strip() or "macOS"
    os_version_str = meta.get("_meta_os_version", "").strip()
    if not os_version_str:
        raise ValueError(
            "Metadata row missing from input file or '_meta_os_version' column not found. "
            "Re-export the baseline to regenerate the metadata row."
        )
    os_version = float(os_version_str)
    os_typeversion = f"{os_type}_{int(os_version)}".lower()

    dry_label = "  [DRY RUN]" if dry_run else ""
    print(f"\nImporting {input_path}  ({os_type} {int(os_version)}){dry_label}\n")

    summary: dict[str, int] = {"modified": 0, "skipped": 0, "not_found": 0, "created": 0}
    categories = _valid_categories()

    for row in rows:
        rule_id = row.get("rule_id", "").strip()
        if not rule_id:
            continue

        rule_file = _find_rule_file(rule_id)
        if rule_file is None:
            # No existing file — create it as a new rule in the custom dir.
            # It must land under a category subfolder (e.g.
            # custom_rules_dir/system_settings/<id>.yaml) matching one of
            # rules_dir's own subfolders -- that's what
            # `collect_platform_rules` scans, and a file dropped directly
            # in custom_rules_dir's root is invisible to it (silently
            # excluded from every baseline).
            title = row.get("title", "").strip()
            category = row.get("category", "").strip()
            if not title:
                print(f"  {'SKIP':<8}  {rule_id}")
                print(f"  {'':<8}    no title — cannot create new rule")
                summary["not_found"] += 1
                continue
            if category not in categories:
                print(f"  {'SKIP':<8}  {rule_id}")
                print(
                    f"  {'':<8}    invalid or missing category "
                    f"{category!r} — must be one of: {', '.join(sorted(categories))}"
                )
                summary["not_found"] += 1
                continue

            populated = [
                col for col in FIXED_COLUMNS
                if col not in ("rule_id", "category") and row.get(col, "").strip()
            ]
            new_yaml_data = _create_rule_yaml(row, os_type, os_version, os_typeversion)
            custom_rules_dir = Path(config["custom"]["rules_dir"]) / category
            new_rule_path = custom_rules_dir / f"{rule_id}.yaml"

            if not dry_run:
                custom_rules_dir.mkdir(parents=True, exist_ok=True)
                with new_rule_path.open("w", encoding="utf-8") as f:
                    _ryaml.dump(new_yaml_data, f)

            print(f"  {'CREATE':<8}  {rule_id}  →  {new_rule_path}")
            if populated:
                print(f"  {'':<8}    {',  '.join(populated)}")
            summary["created"] += 1
            continue

        with rule_file.open("r", encoding="utf-8") as f:
            yaml_data = _ryaml.load(f)

        changed_fields = _apply_row_to_yaml(
            yaml_data, row, os_type, os_version, os_typeversion
        )

        if changed_fields:
            if not dry_run:
                with rule_file.open("w", encoding="utf-8") as f:
                    _ryaml.dump(yaml_data, f)
            print(f"  {'UPDATE':<8}  {rule_id}  →  {rule_file}")
            print(f"  {'':<8}    {',  '.join(changed_fields)}")
            summary["modified"] += 1
        else:
            summary["skipped"] += 1

    # ── summary footer ────────────────────────────────────────────────────────
    parts = []
    if summary["created"]:
        parts.append(f"{summary['created']} created")
    if summary["modified"]:
        parts.append(f"{summary['modified']} updated")
    if summary["skipped"]:
        parts.append(f"{summary['skipped']} unchanged")
    if summary["not_found"]:
        parts.append(f"{summary['not_found']} skipped (no title)")
    print(f"\n  {' · '.join(parts) if parts else 'nothing to do'}\n")

    return summary


# ── CLI entry points ───────────────────────────────────────────────────────────

def generate_spreadsheet_export(args: Any) -> None:
    """CLI handler for ``mscp admin export``."""
    suffix = ".xlsx" if getattr(args, "xlsx", False) else ".csv"
    baseline_arg = getattr(args, "baseline", None)
    if baseline_arg:
        baseline = Baseline.from_yaml(baseline_arg)
        output_path: Path = getattr(args, "output", None) or (
            Path(config["output_dir"]) / f"{baseline.name}{suffix}"
        )
        export_spreadsheet(baseline, Path(output_path))
    else:
        os_name: str = getattr(args, "os_name", "macos") or "macos"
        os_type = os_name.replace("os", "OS")
        os_version = float(getattr(args, "os_version", 26.0) or 26.0)
        output_path = getattr(args, "output", None) or (
            Path(config["output_dir"])
            / f"rule_template_{os_name}_{int(os_version)}{suffix}"
        )
        export_template(os_type, os_version, Path(output_path))


def generate_spreadsheet_import(args: Any) -> None:
    """CLI handler for ``mscp admin import``."""
    dry_run: bool = getattr(args, "dry_run", False)
    import_spreadsheet(Path(args.input), dry_run=dry_run)
